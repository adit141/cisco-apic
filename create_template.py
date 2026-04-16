#!/usr/bin/env python3
"""
=============================================================================
  ACI vPC Deployment – Excel Template Generator
=============================================================================
  Run once to create ACI_VPC_DEPLOY_Template.xlsx with correct column headers
  and two sample rows.

  Usage:
    python create_template.py
=============================================================================
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    "tenant_name",        # Tenant name in ACI
    "app_profile",        # Application Profile name
    "epg_name",           # EPG name
    "vlan_id",            # VLAN encapsulation ID (e.g. 100)
    "pod_id",             # ACI Pod ID (usually 1)
    "leaf1",              # First leaf node ID  (e.g. 101)
    "leaf2",              # Second leaf node ID (e.g. 102)
    "int_profile_leaf1",  # Leaf interface profile name for leaf1
    "int_profile_leaf2",  # Leaf interface profile name for leaf2
    "ipg_name",           # vPC Interface Policy Group name
    "aaep_name",          # Attachable Access Entity Profile name
    "link_speed",         # Link Level Policy name (e.g. HIF-1G, HIF-10G)
    "from_port",          # Starting port number (e.g. 1)
    "to_port",            # Ending port number   (e.g. 1)
    "port_desc",          # Interface selector description (optional)
    "mode",               # EPG binding mode: regular | native | untagged
]

# ── Sample rows ───────────────────────────────────────────────────────────────
SAMPLE_ROWS = [
    {
        "tenant_name":       "TN-PROD",
        "app_profile":       "AP-WEB",
        "epg_name":          "EPG-WEB-VLAN100",
        "vlan_id":           100,
        "pod_id":            1,
        "leaf1":             101,
        "leaf2":             102,
        "int_profile_leaf1": "Leaf101-IntProf",
        "int_profile_leaf2": "Leaf102-IntProf",
        "ipg_name":          "IPG-SERVER01-vPC",
        "aaep_name":         "AAEP-SERVERS",
        "link_speed":        "HIF-1G",
        "from_port":         1,
        "to_port":           1,
        "port_desc":         "vPC to SERVER01 DATA",
        "mode":              "regular",
    },
    {
        "tenant_name":       "TN-PROD",
        "app_profile":       "AP-DB",
        "epg_name":          "EPG-DB-VLAN200",
        "vlan_id":           200,
        "pod_id":            1,
        "leaf1":             101,
        "leaf2":             102,
        "int_profile_leaf1": "Leaf101-IntProf",
        "int_profile_leaf2": "Leaf102-IntProf",
        "ipg_name":          "IPG-SERVER02-vPC",
        "aaep_name":         "AAEP-SERVERS",
        "link_speed":        "HIF-10G",
        "from_port":         2,
        "to_port":           2,
        "port_desc":         "vPC to SERVER02 DB",
        "mode":              "regular",
    },
    {
        "tenant_name":       "TN-STAGING",
        "app_profile":       "AP-APP",
        "epg_name":          "EPG-APP-VLAN300",
        "vlan_id":           300,
        "pod_id":            1,
        "leaf1":             103,
        "leaf2":             104,
        "int_profile_leaf1": "Leaf103-IntProf",
        "int_profile_leaf2": "Leaf104-IntProf",
        "ipg_name":          "IPG-SERVER03-vPC",
        "aaep_name":         "AAEP-STAGING",
        "link_speed":        "HIF-10G",
        "from_port":         3,
        "to_port":           3,
        "port_desc":         "vPC to SERVER03 APP",
        "mode":              "regular",
    },
]

OUTPUT_FILE = "ACI_VPC_DEPLOY_Template.xlsx"

# ── Column documentation (shown in row 2 as light grey hints) ─────────────────
COLUMN_HINTS = {
    "tenant_name":       "e.g. TN-PROD",
    "app_profile":       "e.g. AP-WEB",
    "epg_name":          "e.g. EPG-WEB-VLAN100",
    "vlan_id":           "e.g. 100",
    "pod_id":            "Usually 1",
    "leaf1":             "e.g. 101",
    "leaf2":             "e.g. 102",
    "int_profile_leaf1": "e.g. Leaf101-IntProf",
    "int_profile_leaf2": "e.g. Leaf102-IntProf",
    "ipg_name":          "e.g. IPG-SERVER01-vPC",
    "aaep_name":         "e.g. AAEP-SERVERS",
    "link_speed":        "e.g. HIF-1G or HIF-10G",
    "from_port":         "e.g. 1",
    "to_port":           "e.g. 1",
    "port_desc":         "(optional) free text",
    "mode":              "regular / native / untagged",
}

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

HINT_FILL    = PatternFill("solid", fgColor="D6E4F0")
HINT_FONT    = Font(name="Calibri", italic=True, color="808080", size=9)

DATA_FONT    = Font(name="Calibri", size=10)
ODD_FILL     = PatternFill("solid", fgColor="FFFFFF")
EVEN_FILL    = PatternFill("solid", fgColor="EBF3FB")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")

_THIN   = Side(style="thin",   color="BDD7EE")
_MEDIUM = Side(style="medium", color="1F4E79")
BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_BORDER = Border(
    left=_MEDIUM, right=_MEDIUM, top=_MEDIUM, bottom=_MEDIUM
)

COL_WIDTHS = {
    "tenant_name": 18, "app_profile": 18, "epg_name":          24,
    "vlan_id":     10, "pod_id":       8, "leaf1":               8,
    "leaf2":        8, "int_profile_leaf1": 28, "int_profile_leaf2": 28,
    "ipg_name":    28, "aaep_name":   24, "link_speed":         14,
    "from_port":   12, "to_port":     12, "port_desc":          34,
    "mode":        12,
}


# ── Generator ─────────────────────────────────────────────────────────────────
def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title        = "Deploy"
    ws.freeze_panes = "A3"           # freeze header + hint row
    ws.sheet_view.showGridLines = True

    # ── Row 1: Column headers ─────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell            = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = ALIGN_CENTER
        cell.border     = HEADER_BORDER

    ws.row_dimensions[1].height = 22

    # ── Row 2: Hints / examples ───────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        hint            = COLUMN_HINTS.get(col_name, "")
        cell            = ws.cell(row=2, column=col_idx, value=hint)
        cell.fill       = HINT_FILL
        cell.font       = HINT_FONT
        cell.alignment  = ALIGN_CENTER
        cell.border     = BORDER

    ws.row_dimensions[2].height = 16

    # ── Rows 3+: Sample data ──────────────────────────────────────────────
    for row_idx, record in enumerate(SAMPLE_ROWS, start=3):
        fill = EVEN_FILL if (row_idx % 2 == 0) else ODD_FILL
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value           = record.get(col_name, "")
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font       = DATA_FONT
            cell.fill       = fill
            cell.alignment  = ALIGN_LEFT
            cell.border     = BORDER
        ws.row_dimensions[row_idx].height = 15

    # ── Column widths & auto-filter ───────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            COL_WIDTHS.get(col_name, 16)
        )

    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNS))}{len(SAMPLE_ROWS) + 2}"
    )

    # ── README sheet ──────────────────────────────────────────────────────
    ws_info = wb.create_sheet("README")
    info_lines = [
        ("Cisco ACI – vPC Deployment Template", True),
        ("", False),
        ("KOLOM WAJIB:", True),
        ("  tenant_name        – Nama tenant di ACI", False),
        ("  app_profile        – Nama Application Profile", False),
        ("  epg_name           – Nama EPG", False),
        ("  vlan_id            – VLAN encapsulation ID (angka)", False),
        ("  pod_id             – ACI Pod ID (biasanya 1)", False),
        ("  leaf1              – Node ID leaf pertama (e.g. 101)", False),
        ("  leaf2              – Node ID leaf kedua  (e.g. 102)", False),
        ("  int_profile_leaf1  – Nama interface profile leaf1 di APIC", False),
        ("                       Contoh: Leaf101-IntProf", False),
        ("  int_profile_leaf2  – Nama interface profile leaf2 di APIC", False),
        ("                       Contoh: Leaf102-IntProf", False),
        ("  ipg_name           – Nama Interface Policy Group (vPC)", False),
        ("  aaep_name          – Nama AAEP", False),
        ("  link_speed         – Nama Link Level Policy (e.g. HIF-1G)", False),
        ("  from_port          – Port awal (angka)", False),
        ("  to_port            – Port akhir (angka)", False),
        ("", False),
        ("KOLOM OPSIONAL:", True),
        ("  port_desc     – Deskripsi interface selector (teks bebas)", False),
        ("  mode          – Mode EPG binding: regular / native / untagged", False),
        ("                  Default: regular", False),
        ("", False),
        ("CATATAN:", True),
        ("  • Gunakan sheet 'Deploy' untuk mengisi data.", False),
        ("  • Satu baris = satu vPC deployment.", False),
        ("  • Jalankan script: python aci_deploy.py", False),
        ("  • CDP policy  : hardcoded → CDP-DISABLED", False),
        ("  • LLDP policy : hardcoded → LLDP-ENABLED", False),
        ("  • LACP policy : hardcoded → LACP-ACTIVE", False),
        ("  • protpaths digunakan untuk vPC binding (bukan paths)", False),
    ]

    ws_info.column_dimensions["A"].width = 70
    for row_i, (text, bold) in enumerate(info_lines, start=1):
        cell       = ws_info.cell(row=row_i, column=1, value=text)
        cell.font  = Font(name="Calibri", bold=bold, size=11 if bold else 10)

    wb.save(OUTPUT_FILE)

    print(f"\n  Template berhasil dibuat : {OUTPUT_FILE}")
    print(f"  Sheet                   : Deploy  (data) + README  (panduan)")
    print(f"  Sample rows             : {len(SAMPLE_ROWS)}")
    print(f"\n  Kolom  : {', '.join(COLUMNS)}")
    print("\n  Edit sheet 'Deploy' sesuai kebutuhan, lalu jalankan:")
    print("  python aci_deploy.py\n")


if __name__ == "__main__":
    create_template()
